#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PVJ Extractor — Excel Writer
Part 5/6 of ~3000-line app
Handles structured Excel output with multiple sheets and formatting.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any

from config import OUTPUT_DIR


# ----------------------------------------------------------------------
# 📁 Ensure output directory exists
# ----------------------------------------------------------------------
def ensure_dir(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)


# ----------------------------------------------------------------------
# 🧩 Write styled DataFrame to sheet
# ----------------------------------------------------------------------
def _write_sheet(writer, df: pd.DataFrame, sheet_name: str):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # header style
    header_fill = PatternFill("solid", fgColor="ddebf7")
    header_font = Font(bold=True, color="000000", name="Segoe UI")
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # autosize columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 60)


# ----------------------------------------------------------------------
# 🧾 Write full workbook (Data + Summaries)
# ----------------------------------------------------------------------
def write_full_workbook(
    df: pd.DataFrame,
    summaries: Dict[str, pd.DataFrame],
    out_path: str,
    config: Any = None,
) -> None:
    """
    Create Excel workbook with:
      - Sheet 1: Data (all entries)
      - Sheet 2..n: Summaries (Crop, Applicant_Type)
    """
    ensure_dir(out_path)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _write_sheet(writer, df, "Data")

        for name, sdf in summaries.items():
            _write_sheet(writer, sdf, name[:31])  # Excel sheet name limit

    # final polish
    try:
        wb = load_workbook(out_path)
        ws0 = wb.active
        ws0.freeze_panes = "A2"
        wb.save(out_path)
    except Exception:
        pass


# ----------------------------------------------------------------------
# 📊 Simple test harness
# ----------------------------------------------------------------------
if __name__ == "__main__":
    import pandas as pd
    sample = pd.DataFrame(
        {
            "Reg_No": ["REG/2024/0001", "REG/2024/0002"],
            "Crop": ["Rice", "Cotton"],
            "Applicant": ["ICAR Institute", "Private Seeds Ltd"],
            "Productivity": ["45 q/ha", "30 q/ha"],
        }
    )
    summaries = {
        "Summary_Crop": sample.groupby("Crop").size().reset_index(name="Count")
    }
    out = os.path.join(OUTPUT_DIR, "sample_output.xlsx")
    write_full_workbook(sample, summaries, out)
    print(f"✅ wrote {out}")
